import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side

#LOADING DATA
expense_raw= pd.read_csv('bank_statements.csv')

expense_data = expense_raw[['type', 'mode','amount', 'valueDate', 'narration']]
#print(expense_data)

#OUTPUT TO EXCEL FILE
expense_data.to_excel('expense.xlsx', index = False)

#MODIFYING EXCEL FILE
wb = load_workbook('expense.xlsx')
ws = wb.active

ws.title = 'overall_data'
ws.insert_cols(0)
ws['A1'] = 'S.no'

#MAKING BOLD AND COLOR

thin = Side(border_style='thin', color = "a5a5e6")
double = Side(border_style='double', color = "a5a5e6")

for c in ws[1]:
    c.font = Font(bold=True)
    c.fill=PatternFill("solid", fgColor="CCCCFF")
    c.border = Border(top = double, left = double, right = double, bottom = thin)

#ADDING NUMBERS SERIALLY
for row_num in range(2,ws.max_row+1):
    ws.cell(row=row_num, column=1, value = row_num-1)

#NEW WORKSHEET FOR DEBITS
debit_data=expense_data.loc[expense_data['type']=='DEBIT']
wb.create_sheet('DEBIT')

ws = wb['DEBIT']
for r in dataframe_to_rows(debit_data, index=False, header=True):
    ws.append(r)

ws.insert_cols(0)
ws['A1'] = 'S.no'
for row_num in range(2,ws.max_row+1):
    ws.cell(row=row_num, column=1, value = row_num-1)

#MAKING BOLD AND COLOR
for c in ws[1]:
    c.font = Font(bold=True)
    c.fill=PatternFill("solid", fgColor="CCCCFF")
    c.border = Border(top = double, left = double, right = double, bottom = thin)

#NEW WORKSHEET FOR CREDITS
credit_data=expense_data.loc[expense_data['type']=='CREDIT']
wb.create_sheet('CREDIT')

ws = wb['CREDIT']
for r in dataframe_to_rows(credit_data, index=False, header=True):
    ws.append(r)

ws.insert_cols(0)
ws['A1'] = 'S.no'
for row_num in range(2,ws.max_row+1):
    ws.cell(row=row_num, column=1, value = row_num-1)

#MAKING BOLD AND COLOR
for c in ws[1]:
    c.font = Font(bold=True)
    c.fill=PatternFill("solid", fgColor="CCCCFF")
    c.border = Border(top = double, left = double, right = double, bottom = thin)

#SAVING EXCEL
wb.save('expense.xlsx')

