import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side

def serialNumbers(ws):
    ws.insert_cols(0)
    ws['A1'] = 'S.no'
    for row_num in range(2,ws.max_row+1):
        ws.cell(row=row_num, column=1, value = row_num-1)

def newWorksheet(wb, name, data):
    wb.create_sheet(name)
    sheet = wb[name]
    for r in dataframe_to_rows(data, index=False, header=True):
        sheet.append(r)
    serialNumbers(sheet)
    return sheet

def getMonthlySummary():
    #Load data
    data = pd.read_excel('expense.xlsx', sheet_name='overall_data')

    #Convert valueDate to datetime
    data['valueDate'] = pd.to_datetime(data['valueDate'])

    #Extract month and year
    data['Month'] = data['valueDate'].dt.month

    #Calculate monthly debit and credit summaries
    debit_summary = data[data['type'] == 'DEBIT'].groupby('Month')['amount'].sum().to_dict()
    credit_summary = data[data['type'] == 'CREDIT'].groupby('Month')['amount'].sum().to_dict()

    summary = pd.DataFrame({
        'Debit': debit_summary,
        'Credit': credit_summary
    }).fillna(0)

    summary['Balance'] = summary['Credit'] - summary['Debit']
    return summary

#SPECIAL FUNCTION TO GET SPECIFIC MONTH SUMMARY
def getSpecificMonthSummary(month):
    monthly_summary = getMonthlySummary()
    month = input("Enter month number (1-12) to view summary: ")
    try:
        month_num = int(month)
        if 1 <= month_num <= 12:
            print(f"Monthly Summary for Month {month_num}:")
            print(monthly_summary.loc[month_num])
        else:
            print("Please enter a valid month number between 1 and 12.")
    except ValueError:
        print("Invalid input. Please enter a numeric month number between 1 and 12.")

#AUTO-SIZING COLUMNS SO NOTHING GETS CUT OFF
def autosize(ws):
    for col_cells in ws.columns:
        try:
            max_len = max(
                (len(str(c.value)) 
                 for c in col_cells if c.value is not None),
                default=10,
            )
        except Exception:
            max_len = 10
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)



def styleHeader(ws):
    thin = Side(border_style='thin', color = "a5a5e6")
    double = Side(border_style='double', color = "a5a5e6")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill=PatternFill("solid", fgColor="CCCCFF")
        c.border = Border(top = double, left = double, right = double, bottom = thin)

def Total(data, type):
    return data.loc[data['type'] == type, 'amount'].sum()


def createExcel(input_file, output_file):
    #LOADING DATA
    expense_raw= pd.read_csv(input_file)

    expense_data = expense_raw[['type', 'mode','amount', 'valueDate', 'narration']]

    #MODIFYING EXCEL FILE
    wb = Workbook()
    overall_data = wb.active

    overall_data.title = 'overall_data'

    for row in dataframe_to_rows(expense_data, index = False, header = True):
        overall_data.append(row)
    serialNumbers(overall_data)

    #NEW WORKSHEET FOR DEBITS
    debit_data = expense_data.loc[expense_data['type']=='DEBIT']
    debit_sheet = newWorksheet(wb, 'DEBIT', debit_data)

    #NEW WORKSHEET FOR CREDITS
    credit_data = expense_data.loc[expense_data['type']=='CREDIT']
    credit_sheet = newWorksheet(wb, 'CREDIT', credit_data)

    #TOTAL DEBIT
    debit_sheet.merge_cells('G1:H1')
    debit_sheet.merge_cells('G2:H2')
    total_debit = Total(debit_data, 'DEBIT')
    debit_sheet['G1'] = 'Total Debit'
    debit_sheet['G2'] = total_debit
    debit_sheet['G2'].font = Font(bold=True)

    #TOTAL CREDIT.
    credit_sheet.merge_cells('G1:H1')
    credit_sheet.merge_cells('G2:H2')
    total_credit = Total(credit_data, 'CREDIT')
    credit_sheet['G1'] = 'Total Credit'
    credit_sheet['G2'] = total_credit
    credit_sheet['G2'].font = Font(bold=True)

    #TOTAL BALANCE
    overall_data.merge_cells('G1:H1')
    overall_data.merge_cells('G2:H2')
    balance = total_credit - total_debit
    overall_data['G1'] = 'Balance'
    overall_data['G2'] = balance
    overall_data['G2'].font = Font(bold=True)    


    #MAKING BOLD AND COLOR
    styleHeader(overall_data)
    styleHeader(debit_sheet)
    styleHeader(credit_sheet)

    #Autosize
    for sheet in (overall_data, debit_sheet, credit_sheet):
        autosize(sheet)

    print("createExcel called", flush=True)
    print("Sheets before save:", wb.sheetnames, flush=True)
    print("Output file:", output_file, flush=True)

    #SAVING EXCEL
    wb.save(output_file)

    return output_file





